#!/usr/bin/python3
"""
Author:         Yousif Zito
Purpose:        TPJ655 Capstone Project
App Name:       Advanced Security System
Date:           2022-08-05
"""
# import the necessary packages/libraries
from os import system, environ
from time import *
from datetime import *
from screeninfo import get_monitors
from signal import pause, signal, SIGTERM, SIGINT
from gpiozero import DistanceSensor, Button, LED, Buzzer
from threading import Event
from warnings import filterwarnings
from picamera.array import PiRGBArray
from picamera import PiCamera
from rpi_lcd import LCD
import cv2
from openpyxl import Workbook, load_workbook
from openpyxl.styles import *
from pathlib import Path

exit_event = Event()
# assigning appropriate GPIO pins to devices and declaring global variables.
sensor = DistanceSensor(echo=23, trigger=24)
redLED = LED(19)
greenLED = LED(13)
onCamera = Button(26)
offCamera = Button(21)
buzz = Buzzer(12)
buzzBtn = Button(20)
lcd = LCD()


# Get monitor's ID
monitor = get_monitors()[0]
appName = "Advanced Security System"
print(monitor)
# Grab the Height and Width of the active monitor.
width, height = monitor.width, monitor.height
# initialize the camera and grab a reference to the raw camera capture
camera = PiCamera()
camera.resolution = (width, height)
# camera.resolution = (1080, 600)
camera.framerate = 30 # setting FPS to ~30
rawCapture = PiRGBArray(camera, size=(width, height))
# allow the camera to warmup
sleep(0.1)


# Get current logged username
loggedUserID = environ.get('USER')
print(f"User Logged As: {loggedUserID}")


def excelLogging():
    """
    This function is for Excel sheet.
    It appends a date a time whenever the sensor is triggered.
    """
    # Path for the Excel sheet
    workbook_name = "sensorLogs.xlsx"
    sheet_name = "Sensor Activity Logging"
    excelSheet = Path(f"/home/{loggedUserID}/Desktop/tpj655_code/{workbook_name}")
    print(f"/home/{loggedUserID}/Desktop/tpj655_code/{workbook_name}")

    # Check if Excel sheet exists. If it does then append date and time to it.
    if excelSheet.exists():
        print("Appending Data...")
        # Start by opening the spreadsheet and selecting the main sheet
        workbook = load_workbook(excelSheet)
        # grab the active worksheet
        sheet = workbook.active
        # Set sheet name to Sensor Logging
        sheet.title = sheet_name
        sheet['A1'] = "Sensor Activity"
        sheet['A1'].font = Font(bold=True, size=12)
        sheet.column_dimensions['A'].width = 17
        sheet['B1'] = "Date and Time"
        sheet['B1'].font = Font(bold=True, size=12)
        sheet.column_dimensions['B'].width = 20

        # Append text and date and time when sensor is triggered
        sheet.append(["Sensor Triggered", datetime.now()])
        # Center the data in the cells
        for rows in sheet.iter_rows():
            for cell in rows:
                cell.alignment = Alignment(horizontal="center")

        # # Save the spreadsheet
        # Save the workbook with a
        # filename and close the object
        workbook.save(excelSheet)

    # If does not exist then create a new one called sensorLogs.xlsx
    else:
        print("sensorLogs.xlsx does not exist!\nCreating the excel file & appending data...")
        workbook = Workbook()
        # grab the active worksheet
        sheet = workbook.active
        # Set sheet name to Sensor Logging
        sheet.title = sheet_name
        sheet['A1'] = "Sensor Activity"
        sheet['A1'].font = Font(bold=True, size=12)
        sheet.column_dimensions['A'].width = 17
        sheet['B1'] = "Date and Time"
        sheet['B1'].font = Font(bold=True, size=12)
        sheet.column_dimensions['B'].width = 20

        # Append text and date and time when sensor is triggered
        sheet.append(["Sensor Triggered", datetime.now()])
        # Center the data in the cells
        for rows in sheet.iter_rows():
            for cell in rows:
                cell.alignment = Alignment(horizontal="center")

        # # Save the spreadsheet
        # Save the workbook with a
        # filename and close the object
        workbook.save(excelSheet)


def signal_handler(signum, frame): 
    exit_event.set() 


def exitApp():
    signal(SIGINT, signal_handler)
    sleep(0.5)
    exit(1)


def scrnInitialMsg():
    """
    This function displays initial message on the I2C LCD
    """
    lcd.text(" TPJ655 Project ", 1)
    lcd.text("      2022      ", 2)


def scrnLCD():
    """
    This function prints text on the I2C LCD. And it is called in proximitySensor().
    """
    lcd.clear()
    lcd.text("PRESS THE BUTTON", 1)
    lcd.text("TO RING THE BELL!", 2)


def doorBell():
    """
    This function is to trigger the buzzer whenever the buzzBtn is pressed.
    """
    if buzzBtn.is_pressed:
        sleep(0.3)
        buzz.blink(0.05, 0.05)
    else:
        buzz.off()


def alertSpeaker():
    system('echo "There\'s a person on the door!" | festival --tts')


def buttonCam():
    """
    This function displays the camera on the screen when the Green button is pressed.
    And, pressing Red button will turn OFF the camera. 
    """
    for frame in camera.capture_continuous(rawCapture, format="bgr", use_video_port=True):
        # grab the raw NumPy array representing the image, then initialize the timestamp
        # and occupied/unoccupied text
        image = frame.array
        # show the frame
        cv2.namedWindow(appName, cv2.WND_PROP_FULLSCREEN)
        cv2.moveWindow(appName, monitor.x - 1, monitor.y - 1)
        cv2.setWindowProperty(appName, cv2.WND_PROP_FULLSCREEN, cv2.WINDOW_FULLSCREEN)
        cv2.imshow(appName, image)
        key = cv2.waitKey(1) & 0xFF
        # clear the stream in preparation for the next frame
        rawCapture.truncate(0)
        # Call doorBell() function
        doorBell()
        # if offCamera button is pressed, then break the loop
        if offCamera.is_pressed:
            cv2.destroyAllWindows()
            break


def sensCam():
    """
    This function will display the camera on screen when
    the sensor detects an object or person within specified distance.
    """
    for frame in camera.capture_continuous(rawCapture, format="bgr", use_video_port=True):
        # grab the raw NumPy array representing the image, then initialize the timestamp
        # and occupied/unoccupied text
        image = frame.array
        # show the frame
        cv2.namedWindow(appName, cv2.WND_PROP_FULLSCREEN)
        cv2.moveWindow(appName, monitor.x - 1, monitor.y - 1)
        cv2.setWindowProperty(appName, cv2.WND_PROP_FULLSCREEN, cv2.WINDOW_FULLSCREEN)
        cv2.imshow(appName, image)
        key = cv2.waitKey(1) & 0xFF
        # clear the stream in preparation for the next frame
        rawCapture.truncate(0)
        # Call doorBell() function
        doorBell()
        # if the sensor value greater or equal to 0.2, break from the loop and destroy all opened window frames
        if sensor.value >= 0.2:
            cv2.destroyAllWindows()
            break


def proximitySensor():
    """
    This function checks for the sensor value and based on that,
    it turns ON/OFF either LEDs based on the condition specified.
    Also, it changes the I2C LCD text based on the sensor value too. 
    """
    while True:
        scrnInitialMsg()
        sleep(0.1)
        if sensor.value > 0.2:
            greenLED.on()
        else:
            greenLED.off()
            sleep(0.1)
        if sensor.value > 0.05 and sensor.value < 0.2:
            redLED.on()
            scrnLCD()
            alertSpeaker()
            sensCam()
            excelLogging()
        else:
            redLED.off()
            sleep(0.1)
        if onCamera.is_pressed:
            buttonCam()
        doorBell()


def main():
    try:
        greenLED.on()
        proximitySensor()
    except KeyboardInterrupt:
        exitApp()
        exit(1)

    finally:
        lcd.clear()
        print("Exiting...")


if __name__ == "__main__":
    main()
